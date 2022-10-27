from flask import Flask, render_template, request, redirect, url_for, flash, g, session
from flask_mysqldb import MySQL
import openpyxl



app = Flask(__name__)
app.secret_key = 'many random bytes'

app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'pythonlogin'

mysql = MySQL(app)



class User:
    def __init__(self, id, username, password):
        self.id = id
        self.username = username
        self.password = password

    def __repr__(self):
        return f'<User: {self.username}>'

users = []
users.append(User(id=1, username='admin', password='admin123'))
users.append(User(id=2, username='user', password='user123'))
users.append(User(id=3, username='client', password='client321'))



@app.before_request
def before_request():
    g.user = None

    if 'user_id' in session:
        user = [x for x in users if x.id == session['user_id']][0]
        g.user = user


@app.route('/login', methods=['GET', 'POST'])
def login():
    msg=''
    if request.method == 'POST':
        session.pop('user_id', None)

        username = request.form['username']
        password = request.form['password']
        
        user = [x for x in users if x.username == username][0]
        if user and user.password == password:
            session['user_id'] = user.id
            return redirect(url_for('Index'))
        else:
            msg = 'Incorrect username/password!'
        


    return render_template('login.html',msg=msg)



@app.route('/profile')
def profile():
    if not g.user:
        return redirect(url_for('login'))

    return render_template('profile.html')



    
        

@app.route('/')
def Index():
    cur = mysql.connection.cursor()
    cur.execute("SELECT  * FROM client_info")
    data = cur.fetchall()
    cur.close()

    return render_template('index2.html', students=data )

@app.route('/', methods = ['POST'])
def search():
    if request.method == "POST":
        username = request.form['username']
        cur = mysql.connection.cursor()
        cur.execute('SELECT * FROM client_info WHERE client_name = "'+username+'" OR backend = "'+username+'"')
        data = cur.fetchall()
        cur.close()
        return render_template('index2.html', students = data)



@app.route('/insert', methods = ['POST'])
def insert():

    if request.method == "POST":
        flash("Data Inserted Successfully")
        client_name = request.form['client_name']
        username = request.form['username']
        password = request.form['password']
        upfront = request.form['upfront']
        backend = request.form['backend']
        acc_level = request.form['acc_level']
        contact_info = request.form['contact_info']
        cur = mysql.connection.cursor()
        cur.execute('INSERT INTO client_info VALUES (NULL, %s, %s, %s, %s, %s, %s, %s)', (client_name, username, password, upfront, backend, acc_level, contact_info,))
        mysql.connection.commit()
        return redirect(url_for('Index'))




@app.route('/delete/<string:id_data>', methods = ['GET'])
def delete(id_data):
    flash("Record Has Been Deleted Successfully")
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM client_info WHERE client_id=%s", (id_data,))
    mysql.connection.commit()
    return redirect(url_for('Index'))





@app.route('/update',methods=['POST','GET'])
def update():

    if request.method == 'POST':
        id_data = request.form['id']
        client_name = request.form['client_name']
        username = request.form['username']
        password = request.form['password']
        upfront = request.form['upfront']
        backend = request.form['backend']
        acc_level = request.form['acc_level']
        contact_info = request.form['contact_info']
        cur = mysql.connection.cursor()
        cur.execute("""
               UPDATE client_info
               SET client_name=%s, username=%s, password=%s, upfront=%s, backend=%s, acc_level=%s, contact_info=%s
               WHERE client_id=%s
            """, (client_name, username, password, upfront, backend, acc_level, contact_info, id_data))
        flash("Data Updated Successfully")
        mysql.connection.commit()
        return redirect(url_for('Index'))



@app.route('/upfront')
def upfront():
    cur = mysql.connection.cursor()
    cur.execute("SELECT  * FROM upfront_info")
    data = cur.fetchall()
    cur.close()

    return render_template('upfront.html', data=data )

@app.route('/upfront', methods = ['POST'])
def search2():
    if request.method == "POST":
        username = request.form['username']
        cur = mysql.connection.cursor()
        cur.execute('SELECT * FROM upfront_info WHERE client_name = "'+username+'" OR backend = "'+username+'"')
        data = cur.fetchall()
        cur.close()
        return render_template('upfront.html', data = data)





@app.route('/insert2', methods = ['POST'])
def insert2():

    if request.method == "POST":
        flash("Data Inserted Successfully")
        client_name = request.form['client_name']
        username = request.form['username']
        backend = request.form['backend']
        upfront_given = request.form['upfront_given']
        date_of_purchase = request.form['date_of_purchase']
        amount_received = int(request.form['amount_in']) * int(request.form['dollar_rate'])
        credits_paid = request.form['credits_paid']
        amount_in = request.form['amount_in']
        dollar_rate = request.form['dollar_rate']
        credits_added = request.form['credits_added']
        credits_to_be_added = request.form['credits_to_be_added']
        cur = mysql.connection.cursor()
        cur.execute('INSERT INTO upfront_info VALUES (NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s )', (client_name, username, backend, upfront_given, date_of_purchase, amount_received, credits_paid, amount_in, dollar_rate, credits_added, credits_to_be_added,))
        mysql.connection.commit()
        return redirect(url_for('upfront'))
    
    

@app.route('/delete2/<string:id_data>', methods = ['GET'])
def delete2(id_data):
    flash("Record Has Been Deleted Successfully")
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM upfront_info WHERE id=%s", (id_data,))
    mysql.connection.commit()
    return redirect(url_for('upfront'))


@app.route('/update2',methods=['POST','GET'])
def update2():

    if request.method == 'POST':
        id_data = request.form['id']
        client_name = request.form['client_name']
        username = request.form['username']
        backend = request.form['backend']
        upfront_given = request.form['upfront_given']
        date_of_purchase = request.form['date_of_purchase']
        amount_received = int(request.form['amount_in']) * int(request.form['dollar_rate'])
        credits_paid = request.form['credits_paid']
        amount_in = request.form['amount_in']
        dollar_rate = request.form['dollar_rate']
        credits_added = request.form['credits_added']
        credits_to_be_added = request.form['credits_to_be_added']
        cur = mysql.connection.cursor()
        cur.execute("""
               UPDATE upfront_info
               SET client_name=%s, username=%s, backend=%s, upfront_given=%s, date_of_purchase=%s, amount_received=%s, credits_paid=%s,amount_in=%s, dollar_rate=%s, credits_added=%s, credits_to_be_added=%s
               WHERE id=%s
            """, (client_name, username, backend, upfront_given, date_of_purchase, amount_received, credits_paid, amount_in, dollar_rate, credits_added, credits_to_be_added, id_data))
        flash("Data Updated Successfully")
        mysql.connection.commit()
        return redirect(url_for('upfront'))




@app.route('/add_purchase', methods = ['POST'])
def search3():
    if request.method == "POST":
        username = request.form['username']
        backend = request.form['backend']
        cur = mysql.connection.cursor()
        cur.execute('SELECT upfront FROM client_info WHERE client_name = "'+username+'" AND backend = "'+backend+'"')
        data = cur.fetchone()
        cur.close()
        return render_template('add_purchase.html', data = data)
    
    

@app.route('/download/report/excel')
def download_report():
    
    book = openpyxl.Workbook()
    sheet = book.active
    cur = mysql.connection.cursor()
    # (C) EXPORT DATA TO EXCEL
    cur.execute("SELECT * FROM client_info")
    results = cur.fetchall()
    i = 0
    for row in results:
      i += 1
      j = 1
      for col in row:
        cell = sheet.cell(row = i, column = j)
        cell.value = col
        j += 1
     
    # (D) SAVE EXCEL FILE & CLOSE DB
    filepath = "client_info.xlsx"
    book.save(filepath)
    cur.close()
    return render_template('index2.html' )

@app.route('/upfront/download/report/excel')
def download_report2():
    
    book = openpyxl.Workbook()
    sheet = book.active
    cur = mysql.connection.cursor()
    # (C) EXPORT DATA TO EXCEL
    cur.execute("SELECT * FROM upfront_info")
    results = cur.fetchall()
    i = 0
    for row in results:
      i += 1
      j = 1
      for col in row:
        cell = sheet.cell(row = i, column = j)
        cell.value = col
        j += 1
     
    # (D) SAVE EXCEL FILE & CLOSE DB
    #filepath = '/home/'+getpass.getuser()+'/Desktop/upfront_info.xlsx'
    filepath = 'upfront_info.xlsx'
    book.save(filepath)
    cur.close()
    return render_template('upfront.html' )





@app.route('/logout')
def logout():
    # Remove session data, this will log the user out
   session.pop('loggedin', None)
   session.pop('id', None)
   session.pop('username', None)
   # Redirect to login page
   return redirect(url_for('login'))
    
    





if __name__ == "__main__":
    app.run(debug=True)
